from __future__ import annotations

import io
import json
import os
import re
import shutil
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import streamlit as st

try:
    import anthropic
except ImportError:
    anthropic = None  # type: ignore[assignment, misc]

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    openpyxl = None  # type: ignore[assignment, misc]
    Workbook = None  # type: ignore[assignment, misc]
    Font = None  # type: ignore[assignment, misc]
    DataValidation = None  # type: ignore[assignment, misc]


APP_DIR = Path(__file__).resolve().parent
DEFAULT_JSON_DIR = APP_DIR / "json"
SUPPORTING_DIR = APP_DIR / "supporting_files"
HELP_TXT = SUPPORTING_DIR / "Help.txt"
SLIDES_TEMPLATE_PPTX = SUPPORTING_DIR / "AISEC.Course.Slides Template v0.5.pptx"
CAMTASIA_TEMPLATE_CMPROJ = SUPPORTING_DIR / "Untitled.cmproj"

DEFAULT_CLAUDE_MODEL = os.environ.get("ANTHROPIC_MODEL", "claude-sonnet-4-20250514")
_CLAUDE_AUDIENCE_JSON_HINT = re.compile(
    r"```(?:json)?\s*([\s\S]*?)\s*```", re.IGNORECASE
)


@dataclass
class ValidationIssue:
    level: str  # "error" | "warning"
    message: str


def _load_json(path: Path) -> dict[str, Any]:
    raw = path.read_text(encoding="utf-8")
    data = json.loads(raw)
    if not isinstance(data, dict):
        raise ValueError("Top-level JSON must be an object")
    return data


def _atomic_write_json(path: Path, data: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(json.dumps(data, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    tmp.replace(path)


def _build_script_export_text(data: dict[str, Any]) -> str:
    """Plain text of included scripts only (do_not_include is false)."""
    lessons = data.get("lessons")
    if not isinstance(lessons, list) or not lessons:
        return ""
    lesson0 = lessons[0]
    if not isinstance(lesson0, dict):
        return ""
    scripts = lesson0.get("scripts")
    if not isinstance(scripts, list):
        return ""
    blocks: list[str] = []
    ordered = sorted(scripts, key=lambda d: _coerce_int(d.get("number"), default=0))
    sep = "-" * 72
    slide_export_n = 0
    for sc in ordered:
        if not isinstance(sc, dict):
            continue
        if sc.get("do_not_include"):
            continue
        slide_export_n += 1
        n = _coerce_int(sc.get("number"), default=0)
        body = _as_str(sc.get("script")).rstrip()
        blocks.append(f"Script {n} (Slide {slide_export_n})\n{sep}\n{body}")
    return "\n\n".join(blocks) + ("\n" if blocks else "")


def _build_slides_export_text(data: dict[str, Any]) -> str:
    """Plain text of included slides only (do_not_include is false)."""
    lessons = data.get("lessons")
    if not isinstance(lessons, list) or not lessons:
        return ""
    lesson0 = lessons[0]
    if not isinstance(lesson0, dict):
        return ""
    slides = lesson0.get("slides")
    if not isinstance(slides, list):
        return ""
    blocks: list[str] = []
    ordered = sorted(slides, key=lambda d: _coerce_int(d.get("number"), default=0))
    sep = "-" * 72
    slide_export_n = 0
    for sl in ordered:
        if not isinstance(sl, dict):
            continue
        if sl.get("do_not_include"):
            continue
        slide_export_n += 1
        n = _coerce_int(sl.get("number"), default=0)
        title = _as_str(sl.get("slidetitle")).rstrip()
        subtitle = _as_str(sl.get("subtitle")).rstrip()
        content = _as_str(sl.get("slidecontent")).rstrip()
        body_parts = [f"Title: {title}", f"Subtitle: {subtitle}", f"Content:\n{content}"]
        body = "\n\n".join(body_parts)
        blocks.append(f"Slide {n} (Slide {slide_export_n})\n{sep}\n{body}")
    return "\n\n".join(blocks) + ("\n" if blocks else "")


_PATH_SEGMENT_INVALID = re.compile(r'[/\\:*?"<>|\x00-\x1f]')
_WIN_RESERVED_NAMES = frozenset(
    {"con", "prn", "aux", "nul"}
    | {f"com{i}" for i in range(1, 10)}
    | {f"lpt{i}" for i in range(1, 10)}
)


def _sanitize_category_folder_name(category: Any) -> str:
    """Turn course Category into a single filesystem-safe directory name."""
    raw = _as_str(category).strip()
    if not raw:
        return "Uncategorized"
    name = _PATH_SEGMENT_INVALID.sub("_", raw)
    name = re.sub(r"\s+", " ", name).strip()
    name = name.rstrip(". ")
    if not name or name in (".", ".."):
        return "Uncategorized"
    if name.casefold() in _WIN_RESERVED_NAMES:
        name = f"_{name}_"
    return name


def _package_course_project_dir(json_path: Path, data: dict[str, Any]) -> Path:
    """Directory for a course under (parent of json folder)/package exports/<Category>/<stem>/."""
    stem = json_path.stem
    category_dir = _sanitize_category_folder_name(data.get("Category"))
    package_parent = json_path.parent.parent / "package exports"
    return package_parent / category_dir / stem


# Top-level keys edited on the Course tab (order matches UI).
_COURSE_TAB_FIELD_KEYS: tuple[str, ...] = (
    "Course Title",
    "Category",
    "Who is this for",
    "Team or Dept this is for",
    "Description",
    "Prerequisites",
    "Learning objectives",
    "Duration",
    "Difficulty",
    "Price",
    "Instructor",
    "Published Status",
)


def _course_tab_fields_dict(data: dict[str, Any]) -> dict[str, Any]:
    return {k: data.get(k) for k in _COURSE_TAB_FIELD_KEYS}


def _build_course_title_description_text(data: dict[str, Any]) -> str:
    title = _as_str(data.get("Course Title", "")).rstrip()
    desc = _as_str(data.get("Description", "")).rstrip()
    return f"Course Title:\n{title}\n\nDescription:\n{desc}\n"


def _export_course_info(json_path: Path, data: dict[str, Any]) -> tuple[Path, list[str], list[str]]:
    """Write CourseInfo under the package-exports course folder without overwriting existing files.

    Returns (course_info_dir, written_filenames, skipped_filenames).
    """
    project_dir = _package_course_project_dir(json_path, data)
    info_dir = project_dir / "CourseInfo"
    info_dir.mkdir(parents=True, exist_ok=True)

    written: list[str] = []
    skipped: list[str] = []

    json_path_out = info_dir / "course_info.json"
    if json_path_out.exists():
        skipped.append(json_path_out.name)
    else:
        payload = _course_tab_fields_dict(data)
        json_path_out.write_text(
            json.dumps(payload, indent=2, ensure_ascii=False) + "\n",
            encoding="utf-8",
        )
        written.append(json_path_out.name)

    txt_path_out = info_dir / "course_title_description.txt"
    if txt_path_out.exists():
        skipped.append(txt_path_out.name)
    else:
        txt_path_out.write_text(_build_course_title_description_text(data), encoding="utf-8")
        written.append(txt_path_out.name)

    return info_dir, written, skipped


def _export_course_package(json_path: Path, data: dict[str, Any]) -> Path:
    """Create a folder under (parent of json folder)/package exports/<Category>/<stem>/ with templates, script, and slides text."""
    stem = json_path.stem
    project_dir = _package_course_project_dir(json_path, data)
    if project_dir.exists():
        shutil.rmtree(project_dir)
    project_dir.mkdir(parents=True)
    for name in ("Archives", "Camtasia_Files", "Export", "Script", "Slides"):
        (project_dir / name).mkdir()

    if not HELP_TXT.is_file():
        raise FileNotFoundError(f"Missing supporting file: {HELP_TXT}")
    if not SLIDES_TEMPLATE_PPTX.is_file():
        raise FileNotFoundError(f"Missing supporting file: {SLIDES_TEMPLATE_PPTX}")
    if not CAMTASIA_TEMPLATE_CMPROJ.exists():
        raise FileNotFoundError(f"Missing supporting Camtasia project: {CAMTASIA_TEMPLATE_CMPROJ}")

    shutil.copy2(HELP_TXT, project_dir / "Help.txt")
    slides_dir = project_dir / "Slides"
    shutil.copy2(SLIDES_TEMPLATE_PPTX, slides_dir / f"{stem}.pptx")
    slides_text = _build_slides_export_text(data)
    (slides_dir / f"{stem}.txt").write_text(slides_text, encoding="utf-8")
    dest_cmproj = project_dir / "Camtasia_Files" / f"{stem}.cmproj"
    if CAMTASIA_TEMPLATE_CMPROJ.is_dir():
        shutil.copytree(CAMTASIA_TEMPLATE_CMPROJ, dest_cmproj)
    else:
        shutil.copy2(CAMTASIA_TEMPLATE_CMPROJ, dest_cmproj)

    script_text = _build_script_export_text(data)
    (project_dir / "Script" / f"{stem}.txt").write_text(script_text, encoding="utf-8")
    return project_dir


def _backup_file(path: Path) -> Path:
    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    bak = path.with_suffix(path.suffix + f".{ts}.bak")
    shutil.copy2(path, bak)
    return bak


_CHECKLIST_V1_ROWS: tuple[tuple[str, str], ...] = (
    ("Video", "Slides must not lose animation"),
    ("Video", "Slide flow fits narration as per script"),
    ("Audio", "Must be in stereo"),
    ("Audio", "Remove gaps of silence"),
    ("Audio", "Be between 6db and 12db"),
    ("Audio", "Remove any out of place sounds (clicks, bangs etc) if discovered"),
    ("Video", "Export to respective export folder when editing is complete"),
    ("General", "Ensure all working updated files are stored in the appropriate folder"),
)

_TRACKER_STATUSES: tuple[str, ...] = (
    "Not started",
    "In progress",
    "On hold",
    "Blocked",
    "Complete",
)

_TRACKER_PRIORITIES: tuple[str, ...] = ("Low", "Normal", "High", "Urgent")

_TRACKER_EDIT_STAGES: tuple[str, ...] = (
    "Pre-production",
    "Ready for editing",
    "In edit",
    "Ready for review",
    "Changes requested",
    "Ready for publishing",
    "Published",
)


def _tracker_row_from_course_json(fp: Path) -> tuple[str, str, str]:
    """Returns (video_title, category, notes_suffix). notes_suffix is empty or an error hint."""
    try:
        d = _load_json(fp)
        title = _as_str(d.get("Course Title", "")).strip() or fp.stem
        cat = _as_str(d.get("Category", "")).strip()
        return title, cat, ""
    except Exception as e:
        return fp.stem, "", f"(JSON load error: {e!s})"


def _build_video_production_tracker_xlsx(json_paths: list[Path]) -> bytes:
    """Multi-sheet workbook: production tracker, QC checklist, lookups and guidance."""
    if openpyxl is None or Workbook is None or Font is None or DataValidation is None:
        raise RuntimeError("openpyxl is not installed; run: pip install openpyxl")

    wb = Workbook()
    ws_t = wb.active
    ws_t.title = "Tracker"

    headers = [
        "Video Title",
        "Category",
        "Owner",
        "Status",
        "Priority",
        "Due Date",
        "Edit Stage",
        "Version",
        "Notes",
    ]
    ws_t.append(headers)
    for c in range(1, len(headers) + 1):
        ws_t.cell(row=1, column=c).font = Font(bold=True)

    for fp in json_paths:
        title, category, err_note = _tracker_row_from_course_json(fp)
        ws_t.append([title, category, "", "", "", "", "", "", err_note])

    ws_t.freeze_panes = "A2"
    widths_t = {"A": 44, "B": 36, "C": 18, "D": 22, "E": 12, "F": 14, "G": 22, "H": 10, "I": 48}
    for col, w in widths_t.items():
        ws_t.column_dimensions[col].width = w

    # --- Checklist sheet (tab 2) ---
    ws_c = wb.create_sheet("Checklist")
    ws_c.merge_cells("A1:D1")
    ws_c["A1"] = "Check List v1.0"
    ws_c["A1"].font = Font(bold=True)
    ws_c["A3"] = "Category"
    ws_c["B3"] = "Checklist item"
    ws_c["C3"] = "Done"
    ws_c["D3"] = "Notes"
    for c in range(1, 5):
        ws_c.cell(row=3, column=c).font = Font(bold=True)
    r = 4
    for cat, item in _CHECKLIST_V1_ROWS:
        ws_c.cell(row=r, column=1, value=cat)
        ws_c.cell(row=r, column=2, value=item)
        r += 1
    ws_c.column_dimensions["A"].width = 12
    ws_c.column_dimensions["B"].width = 72
    ws_c.column_dimensions["C"].width = 8
    ws_c.column_dimensions["D"].width = 36

    # --- Lookups sheet (tab 3): documentation + list columns for data validation ---
    ws_l = wb.create_sheet("Lookups")
    ws_l["A1"] = "Tracker column"
    ws_l["B1"] = "Description / how to use"
    doc_rows: list[tuple[str, str]] = [
        ("Video Title", "Course title from JSON (fallback: JSON filename stem). One row per course/video."),
        ("Category", "Copied from the course JSON Category field."),
        ("Owner", "Primary owner for this row (often the video editor)."),
        ("Status", "High-level state — pick from Status values (column F). Use Edit Stage for pipeline step."),
        ("Priority", "Pick from Priority list (column H)."),
        ("Due Date", "Target completion; use a date Excel recognises (YYYY-MM-DD)."),
        ("Edit Stage", "Pipeline step (pre-production → published) — pick from Edit stage values (column J)."),
        ("Version", "e.g. v1, v1.1 after review rounds."),
        ("Notes", "Blockers, links, timestamps, reviewer feedback, export path reminders."),
    ]
    for i, (col_name, desc) in enumerate(doc_rows, start=2):
        ws_l.cell(row=i, column=1, value=col_name)
        ws_l.cell(row=i, column=2, value=desc)

    ws_l["D1"] = "Team role"
    ws_l["E1"] = "Typical responsibilities"
    team_rows: list[tuple[str, str]] = [
        (
            "Reviewer",
            "Watches drafts against script and learning goals; logs change requests and sign-off.",
        ),
        (
            "Course designer",
            "Owns slide content, narration script accuracy, and learning structure; answers content questions.",
        ),
        (
            "Video editor",
            "Builds timeline, audio levels, animations/export; owns technical quality checklist items.",
        ),
    ]
    for i, (role, resp) in enumerate(team_rows, start=2):
        ws_l.cell(row=i, column=4, value=role)
        ws_l.cell(row=i, column=5, value=resp)

    ws_l["F1"] = "Status values"
    for i, v in enumerate(_TRACKER_STATUSES, start=2):
        ws_l.cell(row=i, column=6, value=v)
    status_end = 1 + len(_TRACKER_STATUSES)

    ws_l["H1"] = "Priority values"
    for i, v in enumerate(_TRACKER_PRIORITIES, start=2):
        ws_l.cell(row=i, column=8, value=v)
    pri_end = 1 + len(_TRACKER_PRIORITIES)

    ws_l["J1"] = "Edit stage values"
    for i, v in enumerate(_TRACKER_EDIT_STAGES, start=2):
        ws_l.cell(row=i, column=10, value=v)
    stage_end = 1 + len(_TRACKER_EDIT_STAGES)

    ws_l["A12"] = "Generated"
    ws_l["B12"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws_l["A13"] = "Checklist tab"
    ws_l["B13"] = "Use per video or per export batch; mark Done when each line item is verified."

    for cell in (ws_l["A1"], ws_l["B1"], ws_l["D1"], ws_l["E1"], ws_l["F1"], ws_l["H1"], ws_l["J1"]):
        cell.font = Font(bold=True)

    last_row = max(2, 1 + len(json_paths))
    dv_status = DataValidation(
        type="list",
        formula1=f"=Lookups!$F$2:$F${status_end}",
        allow_blank=True,
    )
    dv_status.error = "Pick a status from the list (see Lookups tab)."
    ws_t.add_data_validation(dv_status)
    dv_status.add(f"D2:D{last_row}")

    dv_pri = DataValidation(
        type="list",
        formula1=f"=Lookups!$H$2:$H${pri_end}",
        allow_blank=True,
    )
    dv_pri.error = "Pick a priority from the list (see Lookups tab)."
    ws_t.add_data_validation(dv_pri)
    dv_pri.add(f"E2:E{last_row}")

    dv_stage = DataValidation(
        type="list",
        formula1=f"=Lookups!$J$2:$J${stage_end}",
        allow_blank=True,
    )
    dv_stage.error = "Pick an edit stage from the list (see Lookups tab)."
    ws_t.add_data_validation(dv_stage)
    dv_stage.add(f"G2:G{last_row}")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _as_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, (int, float, bool)):
        return str(v)
    return str(v)


def _coerce_int(v: Any, *, default: int) -> int:
    if isinstance(v, int):
        return v
    if isinstance(v, float) and v == int(v):
        return int(v)
    if isinstance(v, str) and v.strip().isdigit():
        return int(v.strip())
    return default


def _validate_course(data: dict[str, Any]) -> list[ValidationIssue]:
    issues: list[ValidationIssue] = []

    required_top = [
        "Course Title",
        "Category",
        "Description",
        "Duration",
        "Difficulty",
        "Price",
        "Instructor",
        "Published Status",
        "Prerequisites",
        "Learning objectives",
        "lessons",
    ]
    for k in required_top:
        if k not in data:
            issues.append(ValidationIssue("error", f'Missing required top-level key: "{k}"'))

    lessons = data.get("lessons")
    if lessons is None:
        return issues
    if not isinstance(lessons, list):
        issues.append(ValidationIssue("error", '"lessons" must be a list'))
        return issues
    if not lessons:
        issues.append(ValidationIssue("warning", '"lessons" is empty'))
        return issues

    for idx, lesson in enumerate(lessons, start=1):
        if not isinstance(lesson, dict):
            issues.append(ValidationIssue("error", f"Lesson {idx} must be an object"))
            continue
        for lk in ["lesson title", "order", "Description", "status", "content", "slides", "scripts"]:
            if lk not in lesson:
                issues.append(ValidationIssue("warning", f'Lesson {idx} missing key: "{lk}"'))

        slides = lesson.get("slides")
        if isinstance(slides, list) and slides:
            nums = []
            for s in slides:
                if not isinstance(s, dict):
                    issues.append(ValidationIssue("error", f"Lesson {idx} slide entry must be an object"))
                    continue
                n = s.get("number")
                if not isinstance(n, int):
                    issues.append(ValidationIssue("warning", f"Lesson {idx} slide has non-int number: {n!r}"))
                else:
                    nums.append(n)
                for sk in ["number", "slidetitle", "subtitle", "slidecontent"]:
                    if sk not in s:
                        issues.append(ValidationIssue("warning", f'Lesson {idx} slide missing key: "{sk}"'))
            if nums and len(set(nums)) != len(nums):
                issues.append(ValidationIssue("error", f"Lesson {idx} slides have duplicate numbers"))

        scripts = lesson.get("scripts")
        if isinstance(scripts, list) and scripts:
            nums = []
            for sc in scripts:
                if not isinstance(sc, dict):
                    issues.append(ValidationIssue("error", f"Lesson {idx} script entry must be an object"))
                    continue
                n = sc.get("number")
                if not isinstance(n, int):
                    issues.append(ValidationIssue("warning", f"Lesson {idx} script has non-int number: {n!r}"))
                else:
                    nums.append(n)
                for sk in ["number", "script"]:
                    if sk not in sc:
                        issues.append(ValidationIssue("warning", f'Lesson {idx} script missing key: "{sk}"'))
            if nums and len(set(nums)) != len(nums):
                issues.append(ValidationIssue("error", f"Lesson {idx} scripts have duplicate numbers"))

    return issues


def _sorted_json_files(json_dir: Path) -> list[Path]:
    if not json_dir.exists():
        return []
    return sorted([p for p in json_dir.iterdir() if p.is_file() and p.suffix.lower() == ".json"])


def _lesson0_description(data: dict[str, Any]) -> str:
    lessons = data.get("lessons")
    if not isinstance(lessons, list) or not lessons:
        return ""
    l0 = lessons[0]
    if not isinstance(l0, dict):
        return ""
    return _as_str(l0.get("Description", ""))


def _build_audience_source_text(data: dict[str, Any]) -> str:
    """Plain-text bundle for Claude from course + first-lesson description."""
    chunks = [
        f"Course Title:\n{_as_str(data.get('Course Title'))}",
        f"Category:\n{_as_str(data.get('Category'))}",
        f"Description:\n{_as_str(data.get('Description'))}",
        f"Prerequisites:\n{_as_str(data.get('Prerequisites'))}",
        f"Learning objectives:\n{_as_str(data.get('Learning objectives'))}",
        f"Lesson description (first lesson):\n{_lesson0_description(data)}",
    ]
    return "\n\n".join(chunks)


def _parse_claude_audience_json(raw: str) -> tuple[str, str]:
    """Extract who-is-this-for and team/dept strings from model output."""
    text = raw.strip()
    m = _CLAUDE_AUDIENCE_JSON_HINT.search(text)
    if m:
        text = m.group(1).strip()
    start = text.find("{")
    end = text.rfind("}")
    if start < 0 or end <= start:
        raise ValueError("Model response did not contain a JSON object")
    obj = json.loads(text[start : end + 1])
    if not isinstance(obj, dict):
        raise ValueError("JSON root must be an object")
    who = obj.get("who_is_this_for")
    teams = obj.get("team_or_dept")
    if who is None:
        who = obj.get("Who is this for") or obj.get("whoIsThisFor")
    if teams is None:
        teams = obj.get("Team or Dept this is for") or obj.get("teamOrDept")
    if who is None or teams is None:
        raise ValueError('Expected keys "who_is_this_for" and "team_or_dept" in JSON')
    return _as_str(who).strip(), _as_str(teams).strip()


def _call_claude_for_audience(*, api_key: str, source_text: str, model: str) -> tuple[str, str]:
    if not anthropic:
        raise RuntimeError('Install the anthropic package: pip install anthropic (see review/requirements.txt)')
    system = (
        "You label corporate training courses for a general workforce (non-specialists). "
        "Courses are short awareness-style modules, not aimed at dedicated security engineers—"
        "but you should still name concrete job roles and departments because copy is used for SEO and discovery. "
        "Stay faithful to the supplied content; do not invent unrelated personas or org units."
    )
    user = (
        "From the course metadata below, write two catalog fields.\n\n"
        "1) who_is_this_for: Job roles and personas that benefit (concrete titles people search for). "
        "Comma-separated phrases or short prose is fine.\n"
        "2) team_or_dept: Teams or departments that typically assign or care about this topic; "
        "include recognizable names (e.g. Operations, HR, IT, Legal) when they fit the content.\n\n"
        'Respond with ONLY a JSON object, no markdown fences, using exactly these keys: '
        '"who_is_this_for" and "team_or_dept". Both values must be strings.\n\n'
        "---\n"
        f"{source_text}\n"
        "---"
    )
    client = anthropic.Anthropic(api_key=api_key)
    message = client.messages.create(
        model=model,
        max_tokens=2048,
        messages=[{"role": "user", "content": user}],
        system=system,
    )
    parts: list[str] = []
    for block in message.content:
        if block.type == "text":
            parts.append(block.text)
    combined = "".join(parts)
    return _parse_claude_audience_json(combined)


def _lesson0_slide_script_numbers(lesson0: dict[str, Any]) -> list[int]:
    """Sorted unique slide/script numbers present on the first lesson (union of both lists)."""
    nums: set[int] = set()
    slides = lesson0.get("slides")
    if isinstance(slides, list):
        for sl in slides:
            if isinstance(sl, dict):
                n = _coerce_int(sl.get("number"), default=0)
                if n > 0:
                    nums.add(n)
    scripts = lesson0.get("scripts")
    if isinstance(scripts, list):
        for sc in scripts:
            if isinstance(sc, dict):
                n = _coerce_int(sc.get("number"), default=0)
                if n > 0:
                    nums.add(n)
    return sorted(nums)


def _build_dedup_source_text(data: dict[str, Any]) -> str:
    """Course context plus paired slide fields and scripts for deduplication."""
    lessons = data.get("lessons")
    if not isinstance(lessons, list) or not lessons:
        return ""
    lesson0 = lessons[0]
    if not isinstance(lesson0, dict):
        return ""

    chunks = [
        f"Course Title:\n{_as_str(data.get('Course Title'))}",
        f"Category:\n{_as_str(data.get('Category'))}",
        f"Description:\n{_as_str(data.get('Description'))}",
        f"Prerequisites:\n{_as_str(data.get('Prerequisites'))}",
        f"Learning objectives:\n{_as_str(data.get('Learning objectives'))}",
        f"Lesson title:\n{_as_str(lesson0.get('lesson title'))}",
        f"Lesson description:\n{_as_str(lesson0.get('Description'))}",
        f"Lesson content:\n{_as_str(lesson0.get('content'))}",
    ]

    slides = lesson0.get("slides")
    scripts = lesson0.get("scripts")
    if not isinstance(slides, list):
        slides = []
    if not isinstance(scripts, list):
        scripts = []
    slide_by_n: dict[int, dict[str, Any]] = {}
    for sl in slides:
        if isinstance(sl, dict):
            n = _coerce_int(sl.get("number"), default=0)
            if n:
                slide_by_n[n] = sl
    script_by_n: dict[int, dict[str, Any]] = {}
    for sc in scripts:
        if isinstance(sc, dict):
            n = _coerce_int(sc.get("number"), default=0)
            if n:
                script_by_n[n] = sc

    blocks: list[str] = []
    for n in sorted(set(slide_by_n) | set(script_by_n)):
        sl = slide_by_n.get(n)
        sc = script_by_n.get(n)
        script_body = _as_str(sc.get("script")) if sc else ""
        if sl:
            dni = bool(sl.get("do_not_include"))
            blocks.append(
                f"--- Slide {n} (do_not_include={dni}) ---\n"
                f"slidetitle:\n{_as_str(sl.get('slidetitle'))}\n\n"
                f"subtitle:\n{_as_str(sl.get('subtitle'))}\n\n"
                f"slidecontent:\n{_as_str(sl.get('slidecontent'))}\n\n"
                f"script:\n{script_body}"
            )
        else:
            blocks.append(
                f"--- Slide {n} (slide fields missing in JSON; script only) ---\n"
                f"slidetitle:\n\nsubtitle:\n\nslidecontent:\n\n"
                f"script:\n{script_body}"
            )

    if not blocks:
        return "\n\n".join(chunks)

    return "\n\n".join(chunks) + "\n\n## Slides and scripts (paired by number)\n\n" + "\n\n".join(blocks)


def _parse_claude_dedup_json(raw: str) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    text = raw.strip()
    m = _CLAUDE_AUDIENCE_JSON_HINT.search(text)
    if m:
        text = m.group(1).strip()
    start = text.find("{")
    end = text.rfind("}")
    if start < 0 or end <= start:
        raise ValueError("Model response did not contain a JSON object")
    obj = json.loads(text[start : end + 1])
    if not isinstance(obj, dict):
        raise ValueError("JSON root must be an object")
    slides = obj.get("slides")
    scripts = obj.get("scripts")
    if not isinstance(slides, list) or not isinstance(scripts, list):
        raise ValueError('Expected "slides" and "scripts" arrays in JSON')
    slide_nums: set[int] = set()
    for s in slides:
        if not isinstance(s, dict):
            raise ValueError("Each slide must be an object")
        n = _coerce_int(s.get("number"), default=0)
        if not n:
            raise ValueError("Each slide needs a positive integer number")
        slide_nums.add(n)
    script_nums: set[int] = set()
    for s in scripts:
        if not isinstance(s, dict):
            raise ValueError("Each script must be an object")
        n = _coerce_int(s.get("number"), default=0)
        if not n:
            raise ValueError("Each script needs a positive integer number")
        script_nums.add(n)
    if slide_nums != script_nums:
        raise ValueError("slides and scripts must use the same set of slide numbers")
    return slides, scripts


def _apply_dedup_to_lesson0(
    lesson0: dict[str, Any],
    slides_out: list[dict[str, Any]],
    scripts_out: list[dict[str, Any]],
    *,
    expected_numbers: list[int],
) -> None:
    exp_set = set(expected_numbers)
    slide_nums = {_coerce_int(s.get("number"), default=0) for s in slides_out}
    if slide_nums != exp_set:
        raise ValueError(
            f"Model returned numbers {sorted(slide_nums)}, expected {expected_numbers}"
        )

    slides = lesson0.get("slides")
    if not isinstance(slides, list):
        slides = []
        lesson0["slides"] = slides
    scripts = lesson0.get("scripts")
    if not isinstance(scripts, list):
        scripts = []
        lesson0["scripts"] = scripts

    for s in slides_out:
        if not isinstance(s, dict):
            continue
        n = _coerce_int(s.get("number"), default=0)
        row = _get_by_number(slides, n)
        if row is None:
            raise ValueError(f"No slide with number {n} in course file")
        row["slidetitle"] = _as_str(s.get("slidetitle"))
        row["subtitle"] = _as_str(s.get("subtitle"))
        row["slidecontent"] = _as_str(s.get("slidecontent"))

    for s in scripts_out:
        if not isinstance(s, dict):
            continue
        n = _coerce_int(s.get("number"), default=0)
        row = _get_by_number(scripts, n)
        if row is None:
            raise ValueError(f"No script with number {n} in course file")
        row["script"] = _as_str(s.get("script"))


def _call_claude_for_dedup(
    *, api_key: str, source_text: str, model: str
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    if not anthropic:
        raise RuntimeError('Install the anthropic package: pip install anthropic (see review/requirements.txt)')
    system = (
        "You edit short corporate e-learning courses. Each course has a small fixed set of slide pairs "
        "(on-screen bullets and narration). Your job is to rewrite them so the sequence still teaches the "
        "topic accurately and matches the course metadata, but each slide covers a clearly distinct point: "
        "no duplicated titles, no near-duplicate bullet lists, and no scripts that re-tell the same story "
        "as another slide in paraphrased form. Preserve tone (professional, plain language). "
        "Keep similar length unless tightening removes redundancy. Do not invent unrelated topics. "
        "Do not change which slide numbers exist; return exactly one slide object and one script object per "
        "number you were given. Do not include do_not_include in your JSON (the app keeps existing flags)."
    )
    user = (
        "Analyse the course below. Rewrite slidetitle, subtitle, slidecontent, and script for each slide "
        "number so that together they are faithful to the course but mutually distinct with no obvious overlap.\n\n"
        "Rules:\n"
        "- Each slide should advance the narrative: new angle, example, or takeaway—not a restatement.\n"
        "- Scripts should align with their slide bullets but not repeat other slides' narration.\n"
        "- If two slides currently say the same thing with different words, merge the ideas into one clear "
        "slide and use the other slide numbers for different subtopics implied by the learning objectives.\n"
        '- Respond with ONLY a JSON object (no markdown fences) with keys "slides" and "scripts".\n'
        '- Each is an array of objects with integer "number" matching the input.\n'
        '- Slide objects: number, slidetitle, subtitle, slidecontent (strings).\n'
        '- Script objects: number, script (string).\n\n'
        "---\n"
        f"{source_text}\n"
        "---"
    )
    client = anthropic.Anthropic(api_key=api_key)
    message = client.messages.create(
        model=model,
        max_tokens=16384,
        messages=[{"role": "user", "content": user}],
        system=system,
    )
    parts: list[str] = []
    for block in message.content:
        if block.type == "text":
            parts.append(block.text)
    combined = "".join(parts)
    return _parse_claude_dedup_json(combined)


def _get_by_number(items: list[dict[str, Any]], number: int) -> dict[str, Any] | None:
    for it in items:
        if isinstance(it, dict) and it.get("number") == number:
            return it
    return None


def _ensure_numbered_items(
    items: Any,
    *,
    count: int,
    template: dict[str, Any],
) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    if isinstance(items, list):
        for x in items:
            if isinstance(x, dict):
                out.append(dict(x))
    # ensure 1..count exist
    for n in range(1, count + 1):
        existing = _get_by_number(out, n)
        if existing is None:
            created = dict(template)
            created["number"] = n
            out.append(created)
    out.sort(key=lambda d: _coerce_int(d.get("number"), default=10**9))
    return out


st.set_page_config(page_title="JSON Course Reviewer", layout="wide")

st.title("JSON Course Reviewer")
st.caption("Open, edit, validate, and save course JSON files.")

with st.sidebar:
    st.subheader("Files")
    json_dir_str = st.text_input("JSON folder", value=str(DEFAULT_JSON_DIR))
    json_dir = Path(json_dir_str).expanduser()

    files = _sorted_json_files(json_dir)
    if not files:
        st.warning("No JSON files found in that folder.")
        st.stop()

    file_labels = [p.name for p in files]
    if "file_index" not in st.session_state:
        st.session_state["file_index"] = 0

    # Ensure the stored index is within the current bounds
    st.session_state["file_index"] = min(
        max(st.session_state["file_index"], 0), len(file_labels) - 1
    )

    nav_prev, nav_info, nav_next = st.columns([1, 2, 1])
    with nav_prev:
        if st.button("◀ Prev", use_container_width=True, disabled=st.session_state["file_index"] == 0):
            st.session_state["file_index"] -= 1
    with nav_next:
        if st.button(
            "Next ▶",
            use_container_width=True,
            disabled=st.session_state["file_index"] >= len(file_labels) - 1,
        ):
            st.session_state["file_index"] += 1
    with nav_info:
        st.caption(
            f"File {st.session_state['file_index'] + 1} of {len(file_labels)}"
        )

    jump_to = st.number_input(
        "Jump to file #",
        min_value=1,
        max_value=len(file_labels),
        value=st.session_state["file_index"] + 1,
        step=1,
        key="jump_to_file_number",
    )
    if int(jump_to) - 1 != st.session_state["file_index"]:
        st.session_state["file_index"] = int(jump_to) - 1

    selected_label = st.selectbox(
        "Select a JSON file",
        options=file_labels,
        index=st.session_state["file_index"],
    )
    # Keep the index in sync if the user picks from the dropdown directly
    st.session_state["file_index"] = file_labels.index(selected_label)
    selected_path = next(p for p in files if p.name == selected_label)

    reload_clicked = st.button("Reload from disk")

    st.subheader("AI audience & teams (Claude)")
    st.caption(
        "Fills “Who is this for” and “Team or Dept this is for” from title, category, description, "
        "prereqs, learning objectives, and the first lesson description. Audience is general workforce "
        "(not security specialists), with concrete roles and departments for SEO."
    )
    ai_env_key = (os.environ.get("ANTHROPIC_API_KEY") or "").strip()
    ai_key_input = st.text_input(
        "API key (optional if ANTHROPIC_API_KEY is set)",
        type="password",
        key="claude_api_key_sidebar",
    )
    ai_key = ai_env_key or ai_key_input.strip()
    ai_model = st.text_input("Claude model id", value=DEFAULT_CLAUDE_MODEL, key="claude_model_sidebar")
    ai_backup = st.checkbox("Backup each file (.bak) before write", value=True, key="claude_backup_sidebar")
    r1, r2, r3 = st.columns([1, 1, 2])
    with r1:
        ai_from = st.number_input(
            "From file #",
            min_value=1,
            max_value=len(files),
            value=1,
            help="Start of range in the sorted JSON list (1-based).",
            key="ai_range_from",
        )
    with r2:
        ai_to = st.number_input(
            "To file #",
            min_value=1,
            max_value=len(files),
            value=min(5, len(files)),
            help="End of range (inclusive).",
            key="ai_range_to",
        )
    with r3:
        st.write("")
        deps_ok = anthropic is not None
        if not deps_ok:
            st.caption('Install: pip install "anthropic>=0.25"')
        ai_run = st.button(
            "Fill audience fields (Claude)",
            use_container_width=True,
            disabled=not ai_key or not deps_ok,
            key="ai_batch_run",
        )

    if ai_run and ai_key and deps_ok:
        a, b = int(ai_from), int(ai_to)
        if a > b:
            a, b = b, a
        slice_paths = files[a - 1 : b]
        prog = st.progress(0)
        status = st.empty()
        ok_names: list[str] = []
        fail_msgs: list[str] = []
        model_id = (ai_model or "").strip() or DEFAULT_CLAUDE_MODEL
        total = len(slice_paths)
        for idx, fp in enumerate(slice_paths):
            status.caption(f"Calling Claude for ({idx + 1}/{total}) {fp.name}…")
            prog.progress((idx + 1) / max(total, 1))
            try:
                disk_data = _load_json(fp)
                bundle = _build_audience_source_text(disk_data)
                who, teams = _call_claude_for_audience(
                    api_key=ai_key, source_text=bundle, model=model_id
                )
                disk_data["Who is this for"] = who
                disk_data["Team or Dept this is for"] = teams
                if ai_backup and fp.exists():
                    _backup_file(fp)
                _atomic_write_json(fp, disk_data)
                ok_names.append(fp.name)
            except Exception as e:
                fail_msgs.append(f"{fp.name}: {e!s}")
        status.caption("")
        if ok_names:
            st.success(f"Updated {len(ok_names)} file(s).")
        if fail_msgs:
            st.error("Some files failed:\n\n" + "\n\n".join(fail_msgs))
        if ok_names:
            st.info(
                "If an updated file is open in the editor, click **Reload from disk** to refresh fields."
            )

    st.subheader("Deduplicate slides & scripts (Claude)")
    st.caption(
        "Rewrites slide bullets and narration so each slide advances a distinct idea within the course, "
        "while staying faithful to the topic. Overlap across the ~8 slides is reduced (duplicate or "
        "paraphrased content). Uses the same API key and model as above."
    )
    dedup_backup = st.checkbox(
        "Backup each file (.bak) before write",
        value=True,
        key="dedup_backup_sidebar",
    )
    d1, d2, d3 = st.columns([1, 1, 2])
    with d1:
        dedup_from = st.number_input(
            "From file #",
            min_value=1,
            max_value=len(files),
            value=1,
            help="Start of range in the sorted JSON list (1-based).",
            key="dedup_range_from",
        )
    with d2:
        dedup_to = st.number_input(
            "To file #",
            min_value=1,
            max_value=len(files),
            value=min(5, len(files)),
            help="End of range (inclusive).",
            key="dedup_range_to",
        )
    with d3:
        st.write("")
        dedup_run = st.button(
            "Deduplicate slides/scripts (Claude)",
            use_container_width=True,
            disabled=not ai_key or not deps_ok,
            key="dedup_batch_run",
        )

    if dedup_run and ai_key and deps_ok:
        da, db = int(dedup_from), int(dedup_to)
        if da > db:
            da, db = db, da
        dedup_slice = files[da - 1 : db]
        d_prog = st.progress(0)
        d_status = st.empty()
        d_ok: list[str] = []
        d_fail: list[str] = []
        d_model = (ai_model or "").strip() or DEFAULT_CLAUDE_MODEL
        d_total = len(dedup_slice)
        for d_idx, d_path in enumerate(dedup_slice):
            d_status.caption(f"Deduplicating ({d_idx + 1}/{d_total}) {d_path.name}…")
            d_prog.progress((d_idx + 1) / max(d_total, 1))
            try:
                d_disk = _load_json(d_path)
                lessons_d = d_disk.get("lessons")
                if not isinstance(lessons_d, list) or not lessons_d:
                    raise ValueError("Course has no lessons")
                lesson0_d = lessons_d[0]
                if not isinstance(lesson0_d, dict):
                    raise ValueError("First lesson is invalid")
                expected_d = _lesson0_slide_script_numbers(lesson0_d)
                if not expected_d:
                    raise ValueError("First lesson has no numbered slides")
                bundle_d = _build_dedup_source_text(d_disk)
                if not bundle_d.strip():
                    raise ValueError("Could not build deduplication source text")
                slides_d, scripts_d = _call_claude_for_dedup(
                    api_key=ai_key, source_text=bundle_d, model=d_model
                )
                _apply_dedup_to_lesson0(
                    lesson0_d,
                    slides_d,
                    scripts_d,
                    expected_numbers=expected_d,
                )
                if dedup_backup and d_path.exists():
                    _backup_file(d_path)
                _atomic_write_json(d_path, d_disk)
                d_ok.append(d_path.name)
            except Exception as e:
                d_fail.append(f"{d_path.name}: {e!s}")
        d_status.caption("")
        if d_ok:
            st.success(f"Deduplicated {len(d_ok)} file(s).")
        if d_fail:
            st.error("Some files failed:\n\n" + "\n\n".join(d_fail))
        if d_ok:
            st.info(
                "If an updated file is open in the editor, click **Reload from disk** to refresh fields."
            )


def _load_into_state(path: Path) -> None:
    data = _load_json(path)
    st.session_state["current_path"] = str(path)
    st.session_state["original_data"] = data
    st.session_state["working_data"] = json.loads(json.dumps(data))  # deep copy (json-safe)
    st.session_state["last_loaded_mtime"] = path.stat().st_mtime


need_load = (
    reload_clicked
    or "current_path" not in st.session_state
    or st.session_state.get("current_path") != str(selected_path)
)
if need_load:
    _load_into_state(selected_path)

path = Path(st.session_state["current_path"])
data: dict[str, Any] = st.session_state["working_data"]

top_left, top_right = st.columns([2, 1], gap="large")

export_clicked = False
package_export_clicked = False
course_info_export_clicked = False
course_info_batch_clicked = False
with top_right:
    st.subheader("Validation")
    issues = _validate_course(data)
    errors = [i for i in issues if i.level == "error"]
    warnings = [i for i in issues if i.level == "warning"]
    if not issues:
        st.success("Looks good.")
    else:
        if errors:
            st.error(f"{len(errors)} error(s)")
        if warnings:
            st.warning(f"{len(warnings)} warning(s)")
        with st.expander("Details", expanded=bool(errors)):
            for i in issues:
                (st.error if i.level == "error" else st.warning)(i.message)

    st.subheader("Save")
    make_backup = st.checkbox("Create .bak backup on save", value=True)
    can_save = len(errors) == 0
    save_clicked = st.button("Save to disk", type="primary", disabled=not can_save)
    if not can_save:
        st.caption("Fix validation errors to enable saving.")

    if save_clicked:
        try:
            if make_backup and path.exists():
                bak = _backup_file(path)
                st.info(f"Backup created: {bak.name}")
            _atomic_write_json(path, data)
            st.success("Saved.")
        except Exception as e:
            st.error(f"Save failed: {e!s}")

    export_clicked = st.button("Export Script")
    package_export_clicked = st.button("Export as Package")
    course_info_export_clicked = st.button("Export course info (current file)")
    course_info_batch_clicked = st.button(
        "Export course info (all files in folder)",
        help=(
            "Reads every JSON in the sidebar folder from disk and adds CourseInfo under each "
            "course’s package export path. Skips output files that already exist."
        ),
    )
    st.caption(
        "Course info writes `CourseInfo/course_info.json` and "
        "`CourseInfo/course_title_description.txt` under the package export folder "
        "(existing files are not overwritten)."
    )

    tracker_ok = openpyxl is not None
    if tracker_ok:
        tracker_bytes = _build_video_production_tracker_xlsx(files)
        safe_folder = re.sub(r"[^\w.\-]+", "_", json_dir.name).strip("_")[:48] or "courses"
        tracker_name = f"video_production_tracker_{safe_folder}_{datetime.now():%Y%m%d}.xlsx"
    else:
        tracker_bytes = b""
        tracker_name = "video_production_tracker.xlsx"
    st.download_button(
        label="Download production tracker (Excel)",
        file_name=tracker_name,
        data=tracker_bytes,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        disabled=not tracker_ok,
        help=(
            "Builds a workbook from every JSON in the selected folder: Tracker (titles & categories "
            "from course data), Checklist v1.0, and Lookups with dropdown lists and team roles."
        ),
        key="download_production_tracker_xlsx",
    )
    if not tracker_ok:
        st.caption("Install **openpyxl** to enable the production tracker: `pip install openpyxl`")

    st.subheader("Raw JSON")
    st.download_button(
        label="Download JSON",
        file_name=path.name,
        data=json.dumps(data, indent=2, ensure_ascii=False) + "\n",
        mime="application/json",
    )


with top_left:
    st.subheader(f"Editing: {path.name}")

    tabs = st.tabs(["Course", "Lessons", "Slides", "Scripts", "Preview"])

    with tabs[0]:
        c1, c2 = st.columns(2, gap="large")
        with c1:
            data["Course Title"] = st.text_input("Course Title", value=_as_str(data.get("Course Title", "")))
            data["Category"] = st.text_input("Category", value=_as_str(data.get("Category", "")))
            data["Who is this for"] = st.text_area(
                "Who is this for",
                value=_as_str(data.get("Who is this for", "")),
                height=100,
            )
            data["Team or Dept this is for"] = st.text_area(
                "Team or Dept this is for",
                value=_as_str(data.get("Team or Dept this is for", "")),
                height=100,
            )
            data["Description"] = st.text_area("Description", value=_as_str(data.get("Description", "")), height=120)
            data["Prerequisites"] = st.text_area(
                "Prerequisites", value=_as_str(data.get("Prerequisites", "")), height=220
            )
            data["Learning objectives"] = st.text_area(
                "Learning objectives", value=_as_str(data.get("Learning objectives", "")), height=260
            )

        with c2:
            data["Duration"] = st.text_input("Duration", value=_as_str(data.get("Duration", "")))
            data["Difficulty"] = st.text_input("Difficulty", value=_as_str(data.get("Difficulty", "")))
            data["Price"] = st.text_input("Price", value=_as_str(data.get("Price", "")))
            data["Instructor"] = st.text_input("Instructor", value=_as_str(data.get("Instructor", "")))
            data["Published Status"] = st.text_input(
                "Published Status", value=_as_str(data.get("Published Status", ""))
            )

    lessons = data.get("lessons")
    if not isinstance(lessons, list):
        lessons = []
        data["lessons"] = lessons
    if not lessons:
        lessons.append(
            {
                "lesson title": _as_str(data.get("Course Title", "")),
                "order": 1,
                "Description": "",
                "status": "draft",
                "content": _as_str(data.get("Description", "")),
                "slides": [],
                "scripts": [],
            }
        )

    lesson0 = lessons[0] if isinstance(lessons[0], dict) else {}
    if not isinstance(lesson0, dict):
        lesson0 = {}
        lessons[0] = lesson0

    with tabs[1]:
        lesson0["lesson title"] = st.text_input("Lesson title", value=_as_str(lesson0.get("lesson title", "")))
        lesson0["order"] = _coerce_int(st.text_input("Order", value=_as_str(lesson0.get("order", 1))), default=1)
        lesson0["Description"] = st.text_area(
        "Lesson description", value=_as_str(lesson0.get("Description", "")), height=240
        )
        lesson0["status"] = st.text_input("Lesson status", value=_as_str(lesson0.get("status", "draft")))
        lesson0["content"] = st.text_area("Lesson content", value=_as_str(lesson0.get("content", "")), height=100)

    lesson0_slides = _ensure_numbered_items(
        lesson0.get("slides"),
        count=8,
        template={"slidetitle": "", "subtitle": "", "slidecontent": "", "do_not_include": False},
    )
    lesson0["slides"] = lesson0_slides

    lesson0_scripts = _ensure_numbered_items(
        lesson0.get("scripts"),
        count=8,
        template={"script": "", "do_not_include": False},
    )
    lesson0["scripts"] = lesson0_scripts
    scripts_by_number: dict[int, dict[str, Any]] = {}
    for sc in lesson0_scripts:
        n = _coerce_int(sc.get("number"), default=0)
        if n:
            scripts_by_number[n] = sc

    with tabs[2]:
        st.caption("Edit slide title, subtitle, and bullet content.")
        for slide in lesson0_slides:
            n = _coerce_int(slide.get("number"), default=0)
            with st.expander(f"Slide {n}", expanded=True):
                do_not_include = st.checkbox(
                    "Do Not Include",
                    value=bool(slide.get("do_not_include", False)),
                    key=f"slide_do_not_include_{n}_{path.name}",
                )
                slide["do_not_include"] = bool(do_not_include)
                if n in scripts_by_number:
                    scripts_by_number[n]["do_not_include"] = bool(do_not_include)

                slide["slidetitle"] = st.text_input(
                    f"Slide {n} title",
                    value=_as_str(slide.get("slidetitle", "")),
                    key=f"slide_title_{n}_{path.name}",
                )
                slide["subtitle"] = st.text_input(
                    f"Slide {n} subtitle",
                    value=_as_str(slide.get("subtitle", "")),
                    key=f"slide_subtitle_{n}_{path.name}",
                )
                slide["slidecontent"] = st.text_area(
                    f"Slide {n} content (bullets)",
                    value=_as_str(slide.get("slidecontent", "")),
                    height=160,
                    key=f"slide_content_{n}_{path.name}",
                )

    with tabs[3]:
        st.caption("Edit narration scripts for each slide.")
        for sc in lesson0_scripts:
            n = _coerce_int(sc.get("number"), default=0)
            with st.expander(f"Script {n}", expanded=True):
                st.checkbox(
                    "Do Not Include (from Slide)",
                    value=bool(sc.get("do_not_include", False)),
                    key=f"script_do_not_include_{n}_{path.name}",
                    disabled=True,
                )
                sc["script"] = st.text_area(
                    f"Slide {n} script",
                    value=_as_str(sc.get("script", "")),
                    height=220,
                    key=f"script_{n}_{path.name}",
                )

    with tabs[4]:
        st.json(data, expanded=False)

if export_clicked:
    try:
        export_dir = path.parent / "script exports"
        export_dir.mkdir(parents=True, exist_ok=True)
        out_path = export_dir / path.with_suffix(".txt").name
        text = _build_script_export_text(data)
        out_path.write_text(text, encoding="utf-8")
        st.success(f"Script export saved: {out_path}")
    except Exception as e:
        st.error(f"Export failed: {e!s}")

if package_export_clicked:
    try:
        out_project = _export_course_package(path, data)
        st.success(f"Package export saved: {out_project}")
    except Exception as e:
        st.error(f"Package export failed: {e!s}")

if course_info_export_clicked:
    try:
        info_dir, written, skipped = _export_course_info(path, data)
        lines = [f"Folder: `{info_dir}`"]
        if written:
            lines.append("Wrote: " + ", ".join(written))
        if skipped:
            lines.append("Skipped (already exists): " + ", ".join(skipped))
        msg = "\n\n".join(lines)
        if written:
            st.success(msg)
        else:
            st.info(msg)
    except Exception as e:
        st.error(f"Course info export failed: {e!s}")

if course_info_batch_clicked:
    batch_total = len(files)
    prog = st.progress(0)
    status = st.empty()
    failures: list[str] = []
    courses_ok = 0
    tw = 0
    ts = 0
    for bi, fp in enumerate(files):
        status.caption(f"Course info export ({bi + 1}/{batch_total}): {fp.name}…")
        prog.progress((bi + 1) / max(batch_total, 1))
        try:
            disk = _load_json(fp)
            _info_dir, written, skipped = _export_course_info(fp, disk)
            courses_ok += 1
            tw += len(written)
            ts += len(skipped)
        except Exception as e:
            failures.append(f"{fp.name}: {e!s}")
    status.caption("")
    prog.progress(1.0)
    summary = (
        f"Finished **{batch_total}** JSON file(s): **{courses_ok}** succeeded, "
        f"**{len(failures)}** failed. New files written: **{tw}**; "
        f"skipped (already existed): **{ts}**."
    )
    if failures:
        st.warning(summary)
        with st.expander("Failures", expanded=len(failures) <= 10):
            st.text("\n".join(failures))
    else:
        st.success(summary)

